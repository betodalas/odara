#!/usr/bin/python3

# encoding=utf8
#################################################################################
#  Script     : bank_recon
#  Author     : Farhad Ali
#  Description: This script highlights matching transcations in an excel sheet
#               It requires two sheets to do so:-
#               (i).    usersheet - holds user records
#               (ii).   banksheet - holds records from bank reconciliation report
#################################################################################

##############################################################################
# imports required for the script to work
import openpyxl
from openpyxl import Workbook
from openpyxl.reader import worksheet
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
import sys
import time
import os
import glob
import cgi
import html
import cgitb;

from openpyxl.utils import get_column_letter

cgitb.enable()
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active
# for troubleshooting
from collections import Counter

if sys.version[0] == '2':
    reload(sys)
    sys.setdefaultencoding("utf-8")

# sys.path.insert(0, "/var/www/cgi-bin/bank_recon.py")
os.chdir(".")  # bound path to the current directory

# satisfy openpyxl requirements for highlighting cells
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=12)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill(fill_type='lightUp',
                             start_color='fff000',
                             end_color='6efdfd')

# satisfy openpyxl requirements for highlighting cells2
highlight2 = NamedStyle(name="highlight2")
highlight2.font = Font(bold=True, size=12)
bd = Side(style='thick', color="000000")
highlight2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight2.fill = PatternFill(fill_type='lightUp',
                              start_color='fff000',
                              end_color='fff000')

highlight3 = NamedStyle(name="highlight3")
highlight3.font = Font(bold=True, size=12)
bd = Side(style='thick', color="000000")
highlight3.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight3.fill = PatternFill(fill_type='lightUp',
                             start_color='fff000',
                             end_color='808080')

print("\n")
print("arquivos excel disponiveis nesta pasta:")
print("\n")

# enum files in current directory
# files = os.listdir('/home/roberto/PycharmProjects/odara/')
files = os.listdir('/home/biaenzo/odara/odara')
i = 1
for f in glob.glob("*.xlsx"):
    print("(" + str(i) + "). " + str(f))
    i += 1

print("\n")

print("Content-Type: text/html")  # HTTP header to say HTML is following
print()  # blank line, end of headers

# form = cgi.FieldStorage()
# wb_name  = html.escape(form["wb_name"].value);
# u_sheet  = html.escape(form["u_sheet"].value);
# b_sheet  = html.escape(form["b_sheet"].value);

print("Adicione o nome da planilha em a extensao:")
wb_name = input()
print("\n")
# print(wb_name)
# print(b_sheet)
# print(u_sheet)
time.sleep(1)
try:
    workBook = openpyxl.load_workbook(wb_name + str(".xlsx"))
except IOError:
    print("arquivo nao encontrado")
    sys.exit()
print("Eu encontrei as seguintes abas na planilha")
print("\n")
j = 1;
for sheets in workBook.sheetnames:
    print("(" + str(j) + "). " + str(sheets))
    j += 1
print("\n")
print("Informe o nome da primeira aba")
b_sheet = input()
print("\n")
try:
    bankSheet = workBook[b_sheet]
except KeyError:
    print("Nao encontrei abas nesta planilha " + wb_name + " xlsx.")
    print("exitting....")
    sys.exit()

print("Sucesso, dados encontrados " + b_sheet)
print("\n")
print("Informe o nome da outra aba")
u_sheet = input()
try:
    userSheet = workBook[u_sheet]
    print("\n")
    print("Sucesso, dados encontrados nesta aba: " + u_sheet)
    print("\n")
except KeyError:
    print("Nao encontrei dados nesta aba " + wb_name + " xlsx.")
    sys.exit()

# loop through all records in Column B of the excel file and convert them
# into an array. return that array
# def get_chqs(sheetName):
#    chqs = []
#    for row in range(2, sheetName.max_row + 1):
#        cellObj = sheetName["B" + str(row)]
#        eachChq = cellObj.value
#        chqs.append(eachChq)
#    return chqs

# def get_amount(sheetName):
#    amount = []
#    for row in range(2, sheetName.max_row + 1):
#        cellObj = sheetName["C" + str(row)]
#        eachAmu = cellObj.value
#        amount.append(eachAmu)
#    return amount

# cheques = get_chqs(bankSheet)
# amounts = get_amount(bankSheet)

# print("Numero de verificacoes encontrados " + str(len(cheques)))
# print("Quantidades encontradas " + str(len(amounts)))
# print("\n")

print("Processando ...")
time.sleep(2)
print("Conciliando ...")
time.sleep(2)
count = 0  # keep track of matches found


# for row in range(2, userSheet.max_row + 1):
#    ChcellObject = userSheet["B" + str(row)]        #get cell Object for every record found on column B of excel sheet
#    AmcellObject = userSheet["C" + str(row)]        #same as above for every record on column C of the excel file

#    if ChcellObject.value in cheques:               #check for matches in the "cheque" column
#        if AmcellObject.value in amounts:           #check  for matches in the "amount" column
#            AmcellObject.style = highlight          # highlight them all :)
#            ChcellObject.style = highlight          # --do--
#            count += 1
def get_trios(sheetName):
    trios = []

    def reset(num):
        return str(num).replace('-', '')

    for row in range(2, sheetName.max_row + 1):
        _date = sheetName["A" + str(row)]
        val = sheetName["C" + str(row)]
        val = val.value
        val = reset(val)
        trios.append((_date.value, val))

    return trios


trios = get_trios(sheetName=bankSheet)


def get_trios2(bankSheet2):
    trios2 = []

    def reset(num):
        return str(num).replace('-', '')

    for row in range(2, bankSheet2.max_row + 1):
        _date = bankSheet2["A" + str(row)]
        val = bankSheet2["C" + str(row)]
        val = val.value
        val = reset(val)
        trios2.append((_date.value, val))

    return trios2


trios2 = get_trios2(bankSheet2=userSheet)
print(trios2)


def reset(num):
    return str(num).replace('-', '')


# def get_soma(soma2):
#     compara_soma2 = []

#     for row in range(2, soma2.max_row + 1):
#         _date = soma2["A" + str(row)]
# #        val = soma2["C" = int(row)]
#         compara_soma2.append((_date.value))

#     return compara_soma2

# compara_soma2 = get_soma(soma2=userSheet)

# def get_data(data):
#     datas = []

#     for row in range(2, valor.max_row+ 1):
#         val = valor["C" + str(row)]
# #       val = soma2["C" = int(row)]
#         valores.append((val.value))

#     return valores

# # datas = get_valor(valor=userSheet)


# while True:
#    see_chq = input()
#    if see_chq == "Y":
#        break
#    elif see_chq == "N":
#        print("\n")
#        print("not displaying cheques & amounts..")
#        print("going onward....")
#        time.sleep(2)
#        break
#    else:
#        print("Y or N")


print("processing your data....")
time.sleep(2)
print("finding matches...")
time.sleep(2)
count = 0  # keep track of matches found
bankSheet["E1"].value = 'data'
bankSheet["H1"].value = 'Soma saída'
e_counter = 2

userSheet["D1"].value = 'Entrada'
userSheet["F1"].value = 'Soma valores de saida'
x_counter = 2
date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')

for row in range(2, userSheet.max_row + 1):
    DtcellObject = userSheet["A" + str(row)]
    #    ChcellObject = userSheet["B" + str(row)]        #get cell Object for every record found on column B of excel sheet
    AmcellObject = userSheet["C" + str(row)]
    trio3 = (str(AmcellObject.value))
    campo_valor = AmcellObject.value

    for x in trio3:
        if '-' not in trio3:
            # create new sheet
            DatesObject1 = userSheet["D" + str(x_counter)]
            DatesObject1.value = campo_valor
            x_counter += 1
            AmcellObject.value = None
            AmcellObject.style = highlight3


    # same as above for every record on column C of the excel file

    trio = (DtcellObject.value, reset(AmcellObject.value))

    if trio in trios:
        print('Igual')

        DtcellObject.style = highlight
        #        ChcellObject.style = highlight
        AmcellObject.style = highlight
        print(AmcellObject)
        count += 1

    else:

        DtcellObject.style = highlight2
        #        ChcellObject.style = highlight
        AmcellObject.style = highlight2
        count += 1

        trio3 = (str(AmcellObject.value))

        for x in trio3:
            if '-' not in trio3:
                AmcellObject.style = highlight3

for row in range(2, bankSheet.max_row + 1):

    DtcellObject1 = bankSheet["A" + str(row)]
    #    ChcellObject = userSheet["B" + str(row)]        #get cell Object for every record found on column B of excel sheet
    AmcellObject1 = bankSheet["C" + str(row)]  # same as above for every record on column C of the excel file
    DtcellObject1.style = date_style

    trio2 = (DtcellObject1.value, reset(AmcellObject1.value))

    if trio2 in trios2:
        DtcellObject1.style = highlight
        #        ChcellObject.style = highlight
        AmcellObject1.style = highlight
        count += 1


    else:
        DtcellObject1.style = highlight2
        #        ChcellObject.style = highlight
        AmcellObject1.style = highlight2
        AmcellObject1 = AmcellObject1.value
        print(str(DtcellObject1.value))
        campo_data = DtcellObject1.value
        # book = openpyxl.load_workbook('conciliacao.xlsx')
        # wb = openpyxl.Workbook()

        # sheet = wb.active

        DatesObject = bankSheet["E" + str(e_counter)]
        DatesObject.value = campo_data
        DatesObject.style = date_style
        e_counter += 1

        # wb.save("formula.xlsx")

    #     if AmcellObject.value in amounts:           #check  for matches in the "amount" column
    #         AmcellObject.style = highlight          # highlight them all :)
    #         ChcellObject.style = highlight          # --do--
    #     count += 1

print(str(count) + " matches found")
print("\n")
print("hold on...")
time.sleep(1)
print("highlighting in process...")
time.sleep(2)
print("SUCCESS:" + str(count) + " matches highlighted")
time.sleep(1)
print("creating new file in your folder....")
time.sleep(1)
bankSheet['F2'] = "=SUMIFS($C$2:$C$10000,$A$2:$A$10000,E2)"

bankSheet['H2'] = "=SUM(C2:C1000)"

userSheet['f2'] = "=SUM(C2:C1000)"

workBook.save(
    '/home/biaenzo/odara/odara/conciliacao.xlsx')  # create new file with all the matched instance highlighted automatically
print("conciliacao.xlsx created")
time.sleep(2)
time.sleep(2)
print("Exiting...")

